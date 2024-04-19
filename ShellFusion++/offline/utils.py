import json
import re
import xlsxwriter
from lxml import etree
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import spacy
_spacy = spacy.load('en_core_web_lg')

try:
    import cPickle as pickle
except ImportError:
    import pickle
import os

def posTagger(doc):
    return [[x.text,x.pos_] for x in _spacy(doc)]

def lemmatization(verb):
    return ' '.join([x.lemma_ for x in _spacy(verb)])


def dumpObj(path, obj):
    with open(path, 'wb') as f:
        pickle.dump(obj, f)


def load(path):
    with open(path, 'rb') as f:
        return pickle.load(f)


def writeXlsx(names, lines, xlsx_file):
    """
    Write a list of lines to an xlsx file.
    """
    workbook = xlsxwriter.Workbook(xlsx_file)
    worksheet = workbook.add_worksheet()
    worksheet.write_row(0, 0, names)
    rownum = 1
    for line in lines:
        worksheet.write_row(rownum, 0, line)
        rownum += 1
    workbook.close()


def readXlsx(xlsx_file):
    """
    Read lines from a xlsx file.
    """
    lines = []
    workbook = load_workbook(xlsx_file, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    for i, row in enumerate(worksheet.rows):
        # if i >= 1:
        line = []
        for col in row:
            s = '' if col.value is None else str(col.value).strip()
            line.append(s)
        lines.append(line)
        # line = [ str(col.value).strip() for col in row ]
        # lines.append(line)
    workbook.close()
    return lines


def writeJson(obj, json_file):
    """
    Write an obj to a Json file.
    """
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(obj, f, indent=4, ensure_ascii=False, sort_keys=True)

def readJson(json_file):
    """
    Read an obj from a Json file.
    """
    with open(json_file, 'r', encoding='utf-8') as f:
        obj = json.load(f)
    return obj


def writeDictList2xml(xml_file, dicts, rootname):
    """
    Write a list of dicts (an info unit) to an XML file.
    """
    root = etree.Element(rootname)
    # root = etree.Element(item_name, attrib={'Count': str(len(dict_list))})
    for _dict in dicts:
        try:
            etree.SubElement(root, 'row', attrib=_dict)
        except Exception as e:
            print('***** ERROR in writeDictList2xml():', e, '->', _dict)
    tree = etree.ElementTree(root)
    tree.write(xml_file, pretty_print=True, xml_declaration=True, encoding='utf-8')


def readTxt(txt_file):
    """
    Read the lines of a txt.
    """
    lines = []
    with open(txt_file, 'r', encoding='utf-8') as f:
        for line in f.readlines():
            lines.append(line.strip())
    return lines


def generatePercentageStr(per, acc):
    """
    Generate a percentage str with a specified accuracy.
    """
    if acc == '':
        acc = 2
    acc = '%.' + str(acc) + 'f'
    return  acc % (100 * per) + '%'

def geneImage(xData,yData,R_path):
    xData = list(xData)
    yData = list(yData)
    
    dataLength = len(xData)
    if dataLength != len(yData):return None
    plt.plot(xData,yData)
    plt.imsave(R_path)
    return True

def readJson(path):
    with open(path,'r',encoding='utf-8') as f:
        result = json.load(f)
    return result
def dumpJson(result,path):
    with open(path,'w') as f:
        json.dump(result,f,indent=2)
        
def readXlsx(xlsxPath):
    """
    Read excel table with only one sheet.
    """
    workbook = load_workbook(xlsxPath)
    sheets = workbook.get_sheet_names()
    booksheet = workbook.get_sheet_by_name(sheets[0])
    
    rows = booksheet.rows
    columns = booksheet.columns
    result = list()
    for row in rows:
        result.append([col.value for col in row])
    return result
def outputXls_MultiSheet(xlsx_datas,path,sheets):
    '''
    Generate a excel table with multiple sheets.
    xlsx_data:[{row_name：[row_data]},...]
    sheets:[name0,...]
    '''
    #columnKey = [columnName for columnData in xlsx_data.values() for columnName in columnData.keys()]
    res_xlsx = xlsxwriter.Workbook(path)
    sheet_num = 0 
    for xlsx_data in xlsx_datas:
        #columnKey = list(xlsx_data.values())[0]
        abstract = res_xlsx.add_worksheet(sheets[sheet_num])
        row = 0
        #for i in range(len(columnKey)):
        #    abstract.write(row,i,columnKey[i])
        #row += 1
        for row_name,row_datas in xlsx_data.items():
            row_datas = row_datas
            abstract.write(row,0,row_name)
            for i in range(len(row_datas)):
                abstract.write(row,i+1,row_datas[i])
            row += 1
        sheet_num += 1 
    res_xlsx.close()
def outputXls(xlsx_data,path):
    '''
    Generate a excel table with one sheet.
    xlsx_data:[{row_name：[row_data]},...]
    '''
    res_xlsx = xlsxwriter.Workbook(path)
    abstract = res_xlsx.add_worksheet('abstract')
    row = 0
    for row_name,row_datas in xlsx_data.items():
        abstract.write(row,0,row_name)
        for i in range(len(row_datas)):
            abstract.write(row,i+1,row_datas[i])
        row += 1
    res_xlsx.close()
def ensureInput():
    '''
    Input function for annotation.
    Make sure a number is input during annotation.
    '''
    while True:
        try:
            result = eval(input())
        except SyntaxError:
            print("exception caught.")
            continue
        return result

def readTxt(txt_file):
    """
    Read the lines of a txt.
    """
    lines = []
    with open(txt_file, 'r', encoding='utf-8') as f:
        for line in f.readlines():
            lines.append(line.strip())
    return lines


def readCmdInfo(cmd_info_json):
    """
    Read MP cmds' information.
    """
    cmd_info_dict, cmd_mid_desc_dict = readJson(cmd_info_json), {}

    for cmd in cmd_info_dict:
        mid_desc_dict, mapped_mid = {}, ''
        for mid in cmd_info_dict[cmd]:
            if mid.startswith('man'):
                mid_dict = cmd_info_dict[cmd][mid]
                mp_desc = ' '.join([mid_dict['P-Summary'], mid_dict['P-Option-Description']])
                mid_desc_dict[mid] = { 'MP': mp_desc, 'TLDR': '' }
                if 'TLDR Summary' in mid_dict:
                    mid_desc_dict[mid]['TLDR'] = ' '.join([mid_dict['TLDR P-Summary'], mid_dict['TLDR P-Tasks']])
                    mapped_mid = mid
        cmd_mid_desc_dict[cmd] = { mapped_mid: mid_desc_dict[mapped_mid] } \
            if mapped_mid != '' else mid_desc_dict

    return cmd_info_dict, cmd_mid_desc_dict

def detectOpsInTLDRScript(script):
    """
    Detect the options of a cmd in a TLDR script.
    """
    ops = set()
    for token in script.split()[1:]:
        if token.startswith('-'):
            ops.add(token)
        if re.match('-[a-zA-Z]{2,}$', token):
            for j in range(1, len(token)):
                ops.add('-' + token[j])
    return ops

def generateOpsDes(cmd_ops_dict,cmd_info_dict,cmd_mid_desc_dict):
    #generate mid_dict
    mid_desc_dict = dict()
    cmd_mid_dict = dict()
    for cmdname,ops in cmd_ops_dict.items():
        if cmdname in cmd_mid_desc_dict:
            for mid in cmd_mid_desc_dict[cmdname]:
                mid_desc_dict[mid] = cmd_mid_desc_dict[cmdname][mid]
    result_cmd_op_desc = {mid[mid.find('_')+1:mid.rfind('_')]:{} for mid in mid_desc_dict.keys()}
    for mid in mid_desc_dict.keys():
        cmdname = mid[mid.find('_')+1:mid.rfind('_')]
        cmd_mid_dict[cmdname] = mid

        mid_dict = cmd_info_dict[cmdname][mid]
        mpsumm = mid_dict['Summary']
        result_cmd_op_desc[cmdname]['mpsum'] = mpsumm
        op_desc_dict = mid_dict['Option-Description']
        for op in cmd_ops_dict[cmdname]:
            if op in op_desc_dict.keys():
                result_cmd_op_desc[cmdname][op] = op_desc_dict[op] 
    return result_cmd_op_desc

def generateCmdTLDR(cmd_ops_dict,cmd_mid_desc_dict):
    cmd_tldr_dict = {}
    for cmdname in cmd_ops_dict.keys():
        if cmdname in cmd_mid_desc_dict:
            cmd_tldr_dict[cmdname] = {}
            for mid in cmd_mid_desc_dict[cmdname]:
                cmd_tldr_dict[cmdname][mid] = cmd_mid_desc_dict[cmdname][mid]['TLDR']
    return cmd_tldr_dict

def load(path):
    with open(path, 'rb') as f:
        return pickle.load(f)
    